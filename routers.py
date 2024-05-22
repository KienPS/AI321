from dataclasses import dataclass
import io
from typing import Annotated
import base64

from litestar import get, post, Controller
from litestar.params import Body
from litestar.enums import RequestEncodingType
from litestar.response import Template
from litestar.datastructures import UploadFile, State
from litestar.response import File

from pydantic import BaseModel, BaseConfig
import numpy as np
import cv2
import openpyxl 

import detectors as dt


class TOEICAnswerSheet(BaseModel):
    image: UploadFile
    scanned: bool = False
    
    class Config(BaseConfig):
        arbitrary_types_allowed=True


@dataclass
class TOEICExcelExport(BaseModel):
    type: str

class ToeicParserController(Controller):
    
    @get(path="/", name="toeic_image_upload_form")
    async def get_toeic_image_upload_form(self) -> Template:
        return Template(template_name="toeic_image_form.html.jinja2")
    
    @post(path="/parse", name="parse_toeic_image")
    async def parse_toeic_answer_sheet_image(
            self,
            state: State,
            data: Annotated[TOEICAnswerSheet, Body(media_type=RequestEncodingType.MULTI_PART)]
    ) -> Template:
        content = await data.image.read()
        width, height = 2481, 3508
        image = np.fromstring(content, np.uint8)
        image = cv2.imdecode(image, cv2.IMREAD_COLOR)
        toeic_answer_sheet = cv2.resize(image, (width, height))
        if not data.scanned:
            toeic_answer_sheet = dt.findFullAnswerSheet(toeic_answer_sheet, width, height)
        
        listening_test, reading_test = dt.get_listen_test_image(toeic_answer_sheet), dt.get_reading_test_image(toeic_answer_sheet)
        listening_test_set, reading_test_set = dt.get_columns(listening_test), dt.get_columns(reading_test)
        
        user_listening_answer, user_reading_answer = [], []
        listening_answer_keys, reading_answer_keys = ['A'] * 100, ['B'] * 100 # For demo purpose only
        
        image_clone = toeic_answer_sheet.copy()
        for i, sheet in enumerate(listening_test_set):
            keys = listening_answer_keys[25*i:25*(i+1)]
            ans = dt.get_my_ans(sheet, keys, image_clone, i, 'listening')
            for x in ans:
                user_listening_answer.append(x)
            
        for i, sheet in enumerate(reading_test_set):
            keys = reading_answer_keys[25*i:25*(i+1)]
            ans = dt.get_my_ans(sheet, keys, image_clone, i, 'reading')
            for x in ans:
                user_reading_answer.append(x)
        new_shape = (int(image_clone.shape[1]*0.4), int(image_clone.shape[0]*0.4))
        image_clone = cv2.resize(image_clone, new_shape)
                
        toeic_result = {
            'listening': [],
            'reading': []
        }
        listening_score, reading_score = 0, 0
        
        for i, (user, key) in enumerate(zip(user_listening_answer, listening_answer_keys)):
            toeic_result['listening'].append({'question': i+1, 'expected': key, 'yours': user, 'correct': user==key})
            if user == key:
                listening_score += 1
        
        for i, (user, key) in enumerate(zip(user_reading_answer, reading_answer_keys)):
            toeic_result['reading'].append({'question': i+1, 'expected': key, 'yours': user, 'correct': user==key})
            if user == key:
                reading_score += 1
            
        cv2.putText(image_clone, f"{listening_score}/100", (10, 70), cv2.FONT_HERSHEY_SIMPLEX, 3, (0, 0, 255), 5)
        cv2.putText(image_clone, f"{reading_score}/100", (10, 160), cv2.FONT_HERSHEY_SIMPLEX, 3, (0, 0, 255), 5)

        state.toeic = {
            'toeic_result': toeic_result,
            'listening_score': listening_score,
            'reading_score': reading_score,
        }
        
        return Template(
            template_name='toeic_image_result.html.jinja2',
            context={
                'toeic_result': toeic_result,
                'listening_score': listening_score,
                'reading_score': reading_score,
                'image_base64': base64.b64encode(cv2.imencode('.png', image_clone)[1]).decode()
            }
        )
    
    @post(path="/export", name="export_to_excel")
    async def export_to_excel(self, 
                              state: State, 
                              data: Annotated[TOEICExcelExport, Body(media_type=RequestEncodingType.URL_ENCODED)]
                              ) -> File:
        toeic_result = state.toeic['toeic_result']
        type = data.type.lower()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"TOEIC {type.capitalize()} Result"
        ws.append(['Question', 'Expected', 'Yours', '', 'Score', state.toeic[f'{type}_score']])

        for c in 'ABC':
            ws[f'{c}1'].fill = openpyxl.styles.PatternFill(start_color='808080', end_color='808080', fill_type='solid')

        for question in toeic_result[type]:
            color = '00FF00' if question['correct'] else 'FF0000'
            ws.append([question['question'], question['expected'], question['yours']])
            for c in 'ABC':
                ws[f'{c}{ws.max_row}'].fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
            
        wb.save(f'{type.capitalize()}.xlsx')    
        return File(
            path=f'{type.capitalize()}.xlsx',
            filename=f'{type.capitalize()}.xlsx'
        )
        